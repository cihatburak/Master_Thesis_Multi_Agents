"""
Integrate new product reviews from xlsx files into dataset_final.json.

Steps:
1. Read xlsx files from New_Prod_Reviews/
2. Filter reviews to only matching ASINs  
3. Cap at 15 reviews per product (matching existing dataset convention)
4. Update dataset_final.json with new reviews
5. Remove duplicate products (identified during analysis)

After running this, re-run: python ingest.py
"""

import json
import os
import random
from pathlib import Path

import openpyxl


# Configuration
BASE_DIR = Path(__file__).parent
REVIEWS_DIR = BASE_DIR / "New_Prod_Reviews"
DATASET_PATH = BASE_DIR / "dataset_final.json"
BACKUP_PATH = BASE_DIR / "dataset_final_backup.json"

# Max reviews per product (matching existing dataset convention)
MAX_REVIEWS = 15

# Duplicate products to REMOVE from dataset entirely
DUPLICATE_ASINS_TO_REMOVE = {
    "B0BV8H8HVD",  # ASUS ROG Strix G16 2023 RTX 4060 (duplicate of B0BV7XQ9V9)
    "B0DGQ6X3Q2",  # ASUS ROG Strix G17 G713 32GB (same as B0CTMFZR2J & B0FS16ZHYJ)
    "B0CTMFZR2J",  # ASUS ROG Strix G17 G713 64GB (same model, RAM diff)
    "B0FS16ZHYJ",  # ASUS ROG Strix G17 QHD 32GB  (same model, bundle diff)
}


def parse_xlsx_reviews(filepath: str, target_asin: str) -> list[dict]:
    """Parse an xlsx file and extract reviews matching the target ASIN."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb.active

    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    asin_idx = headers.index("ASIN")
    content_idx = headers.index("Content")
    rating_idx = headers.index("Rating")

    reviews = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_asin = row[asin_idx]
        if row_asin != target_asin:
            continue

        text = row[content_idx]
        rating = row[rating_idx]

        if not text or not isinstance(text, str) or len(text.strip()) < 10:
            continue

        try:
            rating = int(rating)
        except (ValueError, TypeError):
            continue

        reviews.append({
            "asin": target_asin,
            "rating": rating,
            "text": text.strip()
        })

    wb.close()
    return reviews


def sample_balanced_reviews(reviews: list[dict], max_count: int) -> list[dict]:
    """
    Sample reviews with a balanced mix of ratings.
    Tries to get a representative distribution rather than just random.
    """
    if len(reviews) <= max_count:
        return reviews

    # Group by rating
    by_rating = {}
    for r in reviews:
        by_rating.setdefault(r["rating"], []).append(r)

    # Try to get at least 1 from each rating, then fill proportionally
    selected = []
    remaining_slots = max_count

    # First pass: 1 from each available rating
    for rating in sorted(by_rating.keys()):
        if by_rating[rating] and remaining_slots > 0:
            chosen = random.choice(by_rating[rating])
            selected.append(chosen)
            by_rating[rating].remove(chosen)
            remaining_slots -= 1

    # Second pass: fill remaining proportionally
    all_remaining = []
    for rating_reviews in by_rating.values():
        all_remaining.extend(rating_reviews)

    if remaining_slots > 0 and all_remaining:
        random.shuffle(all_remaining)
        selected.extend(all_remaining[:remaining_slots])

    return selected


def main():
    random.seed(42)

    print("=" * 60)
    print("Dataset Integration - New Product Reviews")
    print("=" * 60)

    # Load existing dataset
    print("\n[1/4] Loading existing dataset...")
    with open(DATASET_PATH, "r", encoding="utf-8") as f:
        dataset = json.load(f)
    print(f"      Products in dataset: {len(dataset)}")

    # Backup
    print("\n[2/4] Creating backup...")
    with open(BACKUP_PATH, "w", encoding="utf-8") as f:
        json.dump(dataset, f, indent=4, ensure_ascii=False)
    print(f"      Backup saved to: {BACKUP_PATH}")

    # Remove duplicate ASINs
    print(f"\n[3/4] Removing {len(DUPLICATE_ASINS_TO_REMOVE)} duplicate products...")
    before_count = len(dataset)
    dataset = [p for p in dataset if p["id"] not in DUPLICATE_ASINS_TO_REMOVE]
    removed = before_count - len(dataset)
    print(f"      Removed: {removed} products")
    for asin in DUPLICATE_ASINS_TO_REMOVE:
        print(f"        ❌ {asin}")

    # Process new review files
    print(f"\n[4/4] Processing new review files from {REVIEWS_DIR}...")
    xlsx_files = sorted([f for f in os.listdir(REVIEWS_DIR) if f.endswith(".xlsx")])
    print(f"      Found {len(xlsx_files)} xlsx files")

    products_by_id = {p["id"]: p for p in dataset}
    updated = 0
    total_new_reviews = 0

    for fname in xlsx_files:
        asin = fname.split("-")[0]
        filepath = REVIEWS_DIR / fname

        # Parse reviews from xlsx
        reviews = parse_xlsx_reviews(str(filepath), asin)

        if len(reviews) == 0:
            print(f"\n      ⚠️  {asin}: No matching reviews found, skipping")
            continue

        # Sample balanced reviews
        sampled = sample_balanced_reviews(reviews, MAX_REVIEWS)

        if asin in products_by_id:
            product = products_by_id[asin]
            existing_count = len(product.get("reviews", []))

            # Update reviews
            product["reviews"] = sampled
            product["stats"]["review_count"] = len(sampled)

            print(f"\n      ✅ {asin}: {existing_count} → {len(sampled)} reviews (from {len(reviews)} total)")
            updated += 1
            total_new_reviews += len(sampled)
        else:
            print(f"\n      ⚠️  {asin}: Not found in dataset, skipping")

    # Save updated dataset
    print(f"\n{'=' * 60}")
    print("Saving updated dataset...")
    with open(DATASET_PATH, "w", encoding="utf-8") as f:
        json.dump(dataset, f, indent=4, ensure_ascii=False)

    # Final summary
    products_with_reviews = sum(1 for p in dataset if p.get("reviews") and len(p["reviews"]) > 0)
    total_reviews = sum(len(p.get("reviews", [])) for p in dataset)

    print(f"\n{'=' * 60}")
    print("INTEGRATION COMPLETE")
    print("=" * 60)
    print(f"\n📊 Summary:")
    print(f"   Products updated:        {updated}")
    print(f"   New reviews added:        {total_new_reviews}")
    print(f"   Duplicate products removed: {removed}")
    print(f"\n📦 Dataset now:")
    print(f"   Total products:           {len(dataset)}")
    print(f"   Products with reviews:    {products_with_reviews}")
    print(f"   Products without reviews: {len(dataset) - products_with_reviews}")
    print(f"   Total reviews:            {total_reviews}")
    print(f"\n💾 Saved to: {DATASET_PATH}")
    print(f"\n⚠️  Next step: Run 'python ingest.py' to rebuild ChromaDB!")
    print("=" * 60)


if __name__ == "__main__":
    main()
